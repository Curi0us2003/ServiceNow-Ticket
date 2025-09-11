from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import os
import re
from datetime import datetime
import warnings
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Language detection imports
from langdetect import detect_langs
from langdetect.lang_detect_exception import LangDetectException

# Gemini imports
import requests
import json

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)

# File paths
OPEN_TICKETS_PATH = r"C:\Users\RitabrataRoyChoudhur\OneDrive - GyanSys Inc\Desktop\Python\ServiceNow\open.xlsx"
CLOSED_TICKETS_PATH = r"C:\Users\RitabrataRoyChoudhur\OneDrive - GyanSys Inc\Desktop\Python\ServiceNow\close.xlsx"
DEFAULT_THRESHOLD = 0.75

# Updated required columns to include 'Assigned To' and 'Close Notes'
REQUIRED_COLUMNS = ['Number', 'Short Description', 'Assignment group', 'Customer', 'Created', 'Assigned to']
PREFERRED_COLUMNS = REQUIRED_COLUMNS + ['Close Notes', 'Resolved by']

# Gemini API configuration
GEMINI_API_KEY = "#API_KEY#"
GEMINI_API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"

class GeminiClient:
    """Simple Gemini API client"""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
        self.headers = {"Content-Type": "application/json"}
    
    def generate_content(self, prompt, max_tokens=1500, temperature=0.3):
        """Generate content using Gemini API"""
        payload = {
            "contents": [
                {
                    "parts": [
                        {"text": prompt}
                    ]
                }
            ],
            "generationConfig": {
                "temperature": temperature,
                "maxOutputTokens": max_tokens,
                "candidateCount": 1
            }
        }
        
        try:
            response = requests.post(self.api_url, headers=self.headers, data=json.dumps(payload), timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                if 'candidates' in data and len(data['candidates']) > 0:
                    content = data['candidates'][0]['content']['parts'][0]['text']
                    return content.strip()
                else:
                    raise Exception("No candidates in response")
            else:
                raise Exception(f"API request failed with status {response.status_code}: {response.text}")
                
        except requests.exceptions.Timeout:
            raise Exception("Request timed out")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Network error: {str(e)}")
        except Exception as e:
            raise Exception(f"Gemini API error: {str(e)}")
    
    def test_connection(self):
        """Test the Gemini API connection"""
        try:
            response = self.generate_content("Test connection", max_tokens=10)
            return True, response
        except Exception as e:
            return False, str(e)

# Initialize Gemini client globally
try:
    gemini_api_key = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY)
    gemini_client = GeminiClient(gemini_api_key)
    
    # Test the connection
    is_working, test_result = gemini_client.test_connection()
    if is_working:
        logging.info("Gemini API initialized and tested successfully")
        llm = gemini_client
    else:
        logging.error(f"Gemini API test failed: {test_result}")
        llm = None
        gemini_client = None
        
except Exception as e:
    logging.error(f"Failed to initialize Gemini API: {str(e)}")
    llm = None
    gemini_client = None

# Updated template for generating root cause and suggested fixes separately
analysis_template = """You are a highly experienced SAP IT support specialist specializing in analyzing closed support tickets to provide actionable solutions for similar open tickets.

Using your deep knowledge of SAP systems, modules, and official SAP Notes, consider the following inputs:

Closing Notes from Similar Tickets:
{closing_notes}

Open Ticket Description:
{open_ticket_description}

Please analyze the above information and provide a structured response with TWO distinct sections:

## ROOT CAUSE ANALYSIS:
Identify and explain the likely root causes based on patterns in the closing notes and your SAP expertise. Focus on:
- Technical reasons for the issue
- System configuration problems
- User access or permission issues
- Data inconsistencies
- Integration problems
- Performance bottlenecks

## SUGGESTED RESOLUTION:
Provide a step-by-step solution that an SAP support technician can follow, including:
1. Immediate troubleshooting steps
2. Configuration changes needed
3. Relevant SAP transaction codes and tools
4. SAP Notes or documentation references
5. Testing procedures
6. Preventive measures for future occurrences
7. Escalation procedures if the steps don't resolve the issue

Format your response clearly with these two sections. Use professional technical language suitable for SAP support staff.

If the closing notes lack sufficient detail, provide generalized SAP troubleshooting guidance applicable to this type of issue."""

def detect_text_language(text):
    """
    Detect the language of a given text
    Returns language code or 'unknown' if detection fails
    """
    if pd.isna(text) or str(text).strip() == "":
        return 'unknown'
    
    try:
        # Clean the text for better detection
        cleaned_text = str(text).strip()
        
        # Skip very short texts (less than 3 characters)
        if len(cleaned_text) < 3:
            return 'unknown'
        
        # Detect languages with confidence scores
        lang_probs = detect_langs(cleaned_text)
        
        # Get the most probable language
        if lang_probs:
            most_probable = lang_probs[0]
            # Only return if confidence is reasonable (>0.5)
            if most_probable.prob > 0.5:
                return most_probable.lang
        
        return 'unknown'
        
    except (LangDetectException, Exception) as e:
        logging.debug(f"Language detection failed for text: {str(text)[:50]}... Error: {str(e)}")
        return 'unknown'

def split_short_descriptions(open_df):
    """
    Separate non-English tickets from English tickets based on Short Description
    Returns: (non_english_df, english_df)
    """
    logging.info("Starting language detection for ticket separation...")
    
    # Create a copy to work with
    df = open_df.copy()
    df['Short Description'] = df['Short Description'].fillna("")
    
    # Detect language for each Short Description
    df['DetectedLang'] = df['Short Description'].apply(detect_text_language)
    
    # Define non-English languages we want to filter out
    non_english_langs = ['ja', 'zh-cn', 'zh-tw', 'zh']
    
    # Filter non-English tickets
    non_english_mask = df['DetectedLang'].isin(non_english_langs)
    non_english_df = df[non_english_mask].copy()
    
    # Keep English and unknown language tickets for analysis
    english_df = df[~non_english_mask].copy()
    
    # Log results
    total_tickets = len(df)
    non_english_count = len(non_english_df)
    english_count = len(english_df)
    
    logging.info(f"Language detection results:")
    logging.info(f"  Total tickets: {total_tickets}")
    logging.info(f"  Non-English tickets (excluded): {non_english_count}")
    logging.info(f"  English/Unknown tickets (included): {english_count}")
    
    if non_english_count > 0:
        lang_counts = non_english_df['DetectedLang'].value_counts()
        logging.info(f"  Non-English language breakdown: {dict(lang_counts)}")
    
    return non_english_df, english_df

def clean_value(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

def preprocess_text(text):
    """Enhanced text preprocessing"""
    if pd.isna(text) or text == "":
        return ""
    
    text = str(text).lower()
    # Remove special characters but keep alphanumeric and spaces
    text = re.sub(r'[^a-zA-Z0-9\s]', ' ', text)
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    # Remove common ServiceNow terms that don't add semantic value
    common_terms = ['ticket', 'issue', 'problem', 'error', 'failed', 'failure']
    words = text.split()
    words = [word for word in words if word not in common_terms or len(words) <= 3]
    
    return ' '.join(words).strip()

def generate_root_cause_and_fix_with_gemini(similar_closed_tickets, open_ticket_description=""):
    """
    Generate separate root cause analysis and suggested fix for open ticket 
    using Google Gemini instead of OpenAI
    """
    global gemini_client
    
    try:
        closing_notes = []
        
        # Extract closing notes from similar tickets
        for ticket in similar_closed_tickets:
            closing_note = ticket.get('Close Notes', '').strip()
            if closing_note and closing_note.lower() not in ['n/a', 'na', '', 'none', 'null', 'not applicable']:
                closing_notes.append(f"• {closing_note}")
        
        # If no meaningful closing notes found
        if not closing_notes:
            return {
                'root_cause': "No meaningful closing notes available from similar tickets for root cause analysis.",
                'suggested_fix': "Please investigate manually or contact the appropriate support team for detailed analysis."
            }
        
        # If only one closing note or Gemini client is not available, return basic format
        if len(closing_notes) == 1 or gemini_client is None:
            basic_notes = "\n".join(closing_notes)
            return {
                'root_cause': f"Based on similar resolved ticket:\n\n{basic_notes}\n\n(Manual analysis required for detailed root cause identification)",
                'suggested_fix': f"Follow the resolution approach from similar ticket:\n\n{basic_notes}" + 
                               ("\n\n(Note: AI enhancement unavailable - Gemini client not initialized)" if gemini_client is None else "")
            }
        
        # Use Gemini to generate enhanced analysis when multiple closing notes exist
        closing_notes_text = "\n".join(closing_notes)
        
        # Create the prompt
        prompt = analysis_template.format(
            closing_notes=closing_notes_text,
            open_ticket_description=open_ticket_description
        )
        
        logging.info(f"Generating Gemini-enhanced analysis for ticket with {len(closing_notes)} similar closing notes")
        
        # Generate enhanced analysis using Gemini
        analysis_content = gemini_client.generate_content(
            prompt=prompt,
            max_tokens=1500,
            temperature=0.3
        )
        
        # Split the response into root cause and suggested fix
        root_cause = ""
        suggested_fix = ""
        
        if "## ROOT CAUSE ANALYSIS:" in analysis_content and "## SUGGESTED RESOLUTION:" in analysis_content:
            parts = analysis_content.split("## SUGGESTED RESOLUTION:")
            root_cause_part = parts[0].replace("## ROOT CAUSE ANALYSIS:", "").strip()
            suggested_fix_part = parts[1].strip()
            
            root_cause = root_cause_part if root_cause_part else "Root cause analysis not clearly identified in the response."
            suggested_fix = suggested_fix_part if suggested_fix_part else "Suggested resolution not clearly provided in the response."
        else:
            # If the format is not as expected, treat the whole response as suggested fix
            suggested_fix = analysis_content
            root_cause = "Root cause analysis requires further investigation based on the available information."
        
        logging.info("Gemini-enhanced analysis generated successfully")
        return {
            'root_cause': root_cause,
            'suggested_fix': suggested_fix
        }
        
    except Exception as e:
        logging.error(f"Error generating Gemini-enhanced analysis: {str(e)}")
        
        # Fallback to original method if Gemini fails
        closing_notes = []
        for ticket in similar_closed_tickets:
            closing_note = ticket.get('Close Notes', '').strip()
            if closing_note and closing_note.lower() not in ['n/a', 'na', '']:
                closing_notes.append(f"• {closing_note}")
        
        if closing_notes:
            fallback_notes = "\n".join(closing_notes)
            return {
                'root_cause': f"Based on similar resolved tickets (Gemini analysis failed):\n\n{fallback_notes}\n\nError: {str(e)}",
                'suggested_fix': f"Follow resolution approach from similar tickets:\n\n{fallback_notes}\n\n(Note: AI enhancement failed)"
            }
        else:
            return {
                'root_cause': "No closing notes available for root cause analysis",
                'suggested_fix': "No resolution guidance available from similar tickets"
            }

# Keep the original function name for backward compatibility
def generate_root_cause_and_fix(similar_closed_tickets, open_ticket_description=""):
    """Wrapper function to maintain compatibility"""
    return generate_root_cause_and_fix_with_gemini(similar_closed_tickets, open_ticket_description)

def calculate_semantic_similarity(open_tickets_df, closed_tickets_df, assignment_group_filter=None, threshold=0.75):
    """
    Find similar tickets between open and closed tickets using TF-IDF + Cosine similarity
    with Assignment Group matching and semantic understanding.
    Now only processes English tickets for similarity analysis.
    """
    print("Starting semantic similarity analysis...")
    
    # Validate required columns
    for col in ['Short Description', 'Assignment group']:
        if col not in open_tickets_df.columns:
            raise ValueError(f"'{col}' column not found in open tickets")
        if col not in closed_tickets_df.columns:
            raise ValueError(f"'{col}' column not found in closed tickets")
    
    # Separate English and non-English tickets from open tickets
    try:
        non_english_open_df, english_open_df = split_short_descriptions(open_tickets_df)
    except Exception as e:
        logging.warning(f"Language detection failed: {str(e)}. Processing all tickets.")
        non_english_open_df = pd.DataFrame()
        english_open_df = open_tickets_df.copy()
    
    # Use only English tickets for similarity analysis - RESET INDEX to avoid out of bounds
    open_df = english_open_df.copy().reset_index(drop=True)
    closed_df = closed_tickets_df.copy().reset_index(drop=True)
    
    # Filter by Assignment Group if specified (only filter open tickets)
    if assignment_group_filter and assignment_group_filter != 'All':
        open_df = open_df[open_df['Assignment group'] == assignment_group_filter].reset_index(drop=True)
        
        if open_df.empty:
            print(f"No English open tickets found with 'Assignment group' = '{assignment_group_filter}'")
            return [], {
                'total_open_tickets': len(open_tickets_df),
                'english_open_tickets': len(english_open_df),
                'non_english_open_tickets': len(non_english_open_df),
                'filtered_english_open_tickets': 0
            }
    
    print(f"Total open tickets: {len(open_tickets_df)}")
    print(f"English open tickets: {len(english_open_df)}")
    print(f"Non-English open tickets: {len(non_english_open_df)}")
    print(f"English open tickets for analysis: {len(open_df)}")
    print(f"Closed tickets count: {len(closed_df)}")
    
    # Store language filtering stats
    lang_stats = {
        'total_open_tickets': len(open_tickets_df),
        'english_open_tickets': len(english_open_df),
        'non_english_open_tickets': len(non_english_open_df),
        'filtered_english_open_tickets': len(open_df)
    }
    
    if len(open_df) == 0:
        print("No English open tickets available for analysis")
        return [], lang_stats
    
    # Preprocess texts (no translation needed)
    open_texts = [preprocess_text(desc) for desc in open_df['Short Description'].tolist()]
    closed_texts = [preprocess_text(desc) for desc in closed_df['Short Description'].tolist()]
    
    # Combine all texts for TF-IDF fitting
    all_texts = open_texts + closed_texts
    
    print("Calculating TF-IDF vectors...")
    # Use TF-IDF with n-grams for better semantic understanding
    vectorizer = TfidfVectorizer(
        stop_words='english',
        ngram_range=(1, 3),  # Include trigrams for better context
        max_features=5000,   # Limit features for performance
        min_df=1,           # Minimum document frequency
        max_df=0.95         # Maximum document frequency to filter common terms
    )
    
    try:
        tfidf_matrix = vectorizer.fit_transform(all_texts)
        open_vectors = tfidf_matrix[:len(open_texts)]
        closed_vectors = tfidf_matrix[len(open_texts):]
    except ValueError as e:
        print(f"Error in TF-IDF vectorization: {str(e)}")
        return [], lang_stats
    
    print("Calculating similarity matrix...")
    # Calculate similarity between open and closed tickets
    similarity_matrix = cosine_similarity(open_vectors, closed_vectors)
    
    print(f"Similarity matrix shape: {similarity_matrix.shape}")
    print(f"Open vectors shape: {open_vectors.shape}")
    print(f"Closed vectors shape: {closed_vectors.shape}")
    print(f"Open DF length: {len(open_df)}")
    print(f"Closed DF length: {len(closed_df)}")
    
    results = []
    
    # Use enumerate to get proper array indices instead of DataFrame indices
    for open_idx, (df_idx, open_ticket_row) in enumerate(open_df.iterrows()):
        if open_idx >= similarity_matrix.shape[0]:
            print(f"Warning: Open ticket index {open_idx} exceeds similarity matrix bounds")
            break
            
        open_assignment_group = clean_value(open_ticket_row['Assignment group'])
        
        # Find similar closed tickets with same assignment group
        similar_closed = []
        
        # Use enumerate to get proper array indices for closed tickets too
        for closed_idx, (df_closed_idx, closed_ticket_row) in enumerate(closed_df.iterrows()):
            if closed_idx >= similarity_matrix.shape[1]:
                print(f"Warning: Closed ticket index {closed_idx} exceeds similarity matrix bounds")
                break
                
            closed_assignment_group = clean_value(closed_ticket_row['Assignment group'])
            
            # Check if Assignment Group matches (case-insensitive)
            if open_assignment_group.lower() != closed_assignment_group.lower():
                continue
            
            # Get similarity score using array indices
            try:
                similarity_score = similarity_matrix[open_idx][closed_idx]
            except IndexError as e:
                print(f"IndexError accessing similarity_matrix[{open_idx}][{closed_idx}]: {str(e)}")
                continue
            
            if similarity_score >= threshold:
                # Prepare closed ticket data
                closed_ticket_data = {}
                for col in closed_df.columns:
                    closed_ticket_data[col] = clean_value(closed_ticket_row[col])
                
                closed_ticket_data['Similarity Score'] = round(similarity_score, 3)
                closed_ticket_data['Similarity Percentage'] = f"{similarity_score * 100:.1f}%"
                
                similar_closed.append(closed_ticket_data)
        
        if similar_closed:
            # Sort by similarity score (highest first)
            similar_closed.sort(key=lambda x: x['Similarity Score'], reverse=True)
            
            # Prepare open ticket data
            open_ticket_data = {}
            for col in open_df.columns:
                open_ticket_data[col] = clean_value(open_ticket_row[col])
            
            # Generate root cause and suggested fix with Gemini enhancement
            open_description = clean_value(open_ticket_row.get('Short Description', ''))
            analysis_result = generate_root_cause_and_fix(similar_closed, open_description)
            
            results.append({
                'open_ticket': open_ticket_data,
                'similar_closed_tickets': similar_closed,
                'best_similarity_score': similar_closed[0]['Similarity Score'],
                'total_similar_closed': len(similar_closed),
                'root_cause': analysis_result['root_cause'],
                'suggested_fix': analysis_result['suggested_fix']
            })
    
    # Sort results by best similarity score
    results.sort(key=lambda x: x['best_similarity_score'], reverse=True)
    
    print(f"Found {len(results)} English open tickets with similar closed tickets")
    return results, lang_stats

def get_open_assignment_groups(open_df):
    """Get unique Assignment Groups from open tickets only"""
    if 'Assignment group' in open_df.columns:
        groups = open_df['Assignment group'].dropna().unique().tolist()
        groups = [str(group).strip() for group in groups if str(group).strip()]
        groups.sort()
        return groups
    return []

def validate_columns(df, dataset_name):
    """Validate that required columns are present"""
    missing_required = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    missing_preferred = [col for col in PREFERRED_COLUMNS if col not in df.columns]
    
    return {
        'missing_required': missing_required,
        'missing_preferred': missing_preferred,
        'has_all_required': len(missing_required) == 0
    }

def create_excel_export(results, analysis_params, lang_stats=None):
    """
    Create Excel file with analysis results including separate root cause and suggested fix
    Now includes language filtering statistics
    """
    wb = Workbook()
    
    # Remove default worksheet
    wb.remove(wb.active)
    
    # Create Summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    open_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    closed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary data with language filtering info
    summary_data = [
        ["Analysis Summary", ""],
        ["Assignment Group Filter", analysis_params.get('assignment_group_filter', 'All')],
        ["Similarity Threshold", analysis_params.get('threshold_percentage', 'N/A')],
        ["", ""],
        ["Language Filtering Results", ""],
        ["Total Open Tickets", lang_stats.get('total_open_tickets', 0) if lang_stats else analysis_params.get('total_open_tickets', 0)],
        ["English Open Tickets", lang_stats.get('english_open_tickets', 0) if lang_stats else "N/A"],
        ["Non-English Open Tickets (Excluded)", lang_stats.get('non_english_open_tickets', 0) if lang_stats else "N/A"],
        ["English Tickets Analyzed", lang_stats.get('filtered_english_open_tickets', 0) if lang_stats else analysis_params.get('filtered_open_tickets', 0)],
        ["", ""],
        ["Analysis Results", ""],
        ["English Tickets with Matches", analysis_params.get('open_tickets_with_matches', 0)],
        ["Total Similar Closed Tickets Found", analysis_params.get('total_matches', 0)],
        ["", ""],
        ["System Information", ""],
        ["LLM Status", "Google Gemini Available" if llm is not None else "Unavailable"],
        ["Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    # Write summary
    for row_idx, (key, value) in enumerate(summary_data, 1):
        cell_key = summary_ws.cell(row=row_idx, column=1, value=key)
        cell_value = summary_ws.cell(row=row_idx, column=2, value=value)
        
        if row_idx == 1 or key in ["Language Filtering Results", "Analysis Results", "System Information"]:  # Header rows
            cell_key.font = header_font
            cell_key.fill = header_fill
            cell_value.font = header_font
            cell_value.fill = header_fill
        
        cell_key.border = border
        cell_value.border = border
    
    summary_ws.column_dimensions['A'].width = 35
    summary_ws.column_dimensions['B'].width = 40
    
    # Create Detailed Results sheet
    details_ws = wb.create_sheet("Detailed Results")
    
    # Prepare detailed data
    detailed_data = []
    
    # Headers
    headers = [
        "Match #", "Open Ticket Number", "Open Description", "Open Assignment Group", 
        "Open Customer", "Open Created", "Open Assigned To", "Closed Ticket Number", 
        "Closed Description", "Closed Assignment Group", "Closed Resolved By", 
        "Closed Assigned To", "Close Notes", "Similarity Score", "Similarity %"
    ]
    
    detailed_data.append(headers)
    
    # Data rows
    for match_idx, match in enumerate(results, 1):
        open_ticket = match['open_ticket']
        
        for closed_ticket in match['similar_closed_tickets']:
            row = [
                match_idx,
                open_ticket.get('Number', ''),
                open_ticket.get('Short Description', ''),
                open_ticket.get('Assignment group', ''),
                open_ticket.get('Customer', ''),
                open_ticket.get('Created', ''),
                open_ticket.get('Assigned to', ''),
                closed_ticket.get('Number', ''),
                closed_ticket.get('Short Description', ''),
                closed_ticket.get('Assignment group', ''),
                closed_ticket.get('Resolved by', ''),
                closed_ticket.get('Assigned to', ''),
                closed_ticket.get('Close Notes', ''),
                closed_ticket.get('Similarity Score', ''),
                closed_ticket.get('Similarity Percentage', '')
            ]
            detailed_data.append(row)
    
    # Write detailed data
    for row_idx, row_data in enumerate(detailed_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = details_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            elif col_idx <= 7:  # Open ticket columns
                cell.fill = open_fill
            else:  # Closed ticket columns
                cell.fill = closed_fill
    
    # Adjust column widths
    for col in details_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        details_ws.column_dimensions[column].width = adjusted_width
    
    # Create Root Cause Analysis sheet
    root_cause_ws = wb.create_sheet("Root Cause Analysis")
    
    root_cause_data = [["Open Ticket Number", "Root Cause Analysis"]]
    
    for match in results:
        open_ticket = match['open_ticket']
        root_cause = match['root_cause']
        
        root_cause_data.append([
            open_ticket.get('Number', ''),
            root_cause
        ])
    
    # Write root cause data
    for row_idx, row_data in enumerate(root_cause_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = root_cause_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    root_cause_ws.column_dimensions['A'].width = 20
    root_cause_ws.column_dimensions['B'].width = 80
    
    # Create Suggested Fixes sheet
    fixes_ws = wb.create_sheet("Suggested Fixes")
    
    fixes_data = [["Open Ticket Number", "Suggested Fix"]]
    
    for match in results:
        open_ticket = match['open_ticket']
        suggested_fix = match['suggested_fix']
        
        fixes_data.append([
            open_ticket.get('Number', ''),
            suggested_fix
        ])
    
    # Write fixes data
    for row_idx, row_data in enumerate(fixes_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = fixes_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
    
    fixes_ws.column_dimensions['A'].width = 20
    fixes_ws.column_dimensions['B'].width = 80
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

@app.route('/')
def index():
    try:
        # Check if files exist
        if not os.path.exists(OPEN_TICKETS_PATH):
            return render_template('index.html', error=f"Open tickets file not found: {OPEN_TICKETS_PATH}")
        
        if not os.path.exists(CLOSED_TICKETS_PATH):
            return render_template('index.html', error=f"Closed tickets file not found: {CLOSED_TICKETS_PATH}")
        
        # Load datasets
        open_df = pd.read_excel(OPEN_TICKETS_PATH)
        closed_df = pd.read_excel(CLOSED_TICKETS_PATH)
        
        # Get language statistics for open tickets
        try:
            non_english_open_df, english_open_df = split_short_descriptions(open_df)
            lang_detection_available = True
        except ImportError:
            # If langdetect is not available, use all tickets
            logging.warning("langdetect not available. All tickets will be processed.")
            non_english_open_df = pd.DataFrame()
            english_open_df = open_df
            lang_detection_available = False
        except Exception as e:
            logging.error(f"Error in language detection: {str(e)}")
            non_english_open_df = pd.DataFrame()
            english_open_df = open_df
            lang_detection_available = False
        
        # Validate columns
        open_validation = validate_columns(open_df, 'open tickets')
        closed_validation = validate_columns(closed_df, 'closed tickets')
        
        # Get Assignment Groups from English open tickets only
        assignment_groups = get_open_assignment_groups(english_open_df)
        
        file_info = {
            'open_tickets_count': len(open_df),
            'closed_tickets_count': len(closed_df),
            'english_open_tickets_count': len(english_open_df),
            'non_english_open_tickets_count': len(non_english_open_df),
            'lang_detection_available': lang_detection_available,
            'open_columns': open_df.columns.tolist(),
            'closed_columns': closed_df.columns.tolist(),
            'required_columns': REQUIRED_COLUMNS,
            'preferred_columns': PREFERRED_COLUMNS,
            'missing_open_required': open_validation['missing_required'],
            'missing_closed_required': closed_validation['missing_required'],
            'missing_open_preferred': open_validation['missing_preferred'],
            'missing_closed_preferred': closed_validation['missing_preferred'],
            'assignment_groups': assignment_groups,
            'has_closing_note': 'Close Notes' in closed_df.columns,
            'has_resolved_by': 'Resolved by' in closed_df.columns,
            'llm_available': llm is not None,
            'llm_model': 'Google Gemini 2.0 Flash' if llm is not None else 'None'
        }
        
        # Check if we have minimum required columns for analysis
        if not open_validation['has_all_required']:
            error_msg = f"Open tickets missing required columns: {', '.join(open_validation['missing_required'])}"
            return render_template('index.html', error=error_msg, file_info=file_info)
        
        if not closed_validation['has_all_required']:
            error_msg = f"Closed tickets missing required columns: {', '.join(closed_validation['missing_required'])}"
            return render_template('index.html', error=error_msg, file_info=file_info)
        
        return render_template('index.html', file_info=file_info)
        
    except Exception as e:
        return render_template('index.html', error=str(e))

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        # Check files exist
        if not os.path.exists(OPEN_TICKETS_PATH):
            return jsonify({'success': False, 'error': 'Open tickets file not found'}), 404
        
        if not os.path.exists(CLOSED_TICKETS_PATH):
            return jsonify({'success': False, 'error': 'Closed tickets file not found'}), 404
        
        # Load datasets
        open_df = pd.read_excel(OPEN_TICKETS_PATH)
        closed_df = pd.read_excel(CLOSED_TICKETS_PATH)
        
        # Validate required columns
        for col in ['Short Description', 'Assignment group']:
            if col not in open_df.columns:
                return jsonify({'success': False, 'error': f"'{col}' column missing from open tickets"}), 400
            if col not in closed_df.columns:
                return jsonify({'success': False, 'error': f"'{col}' column missing from closed tickets"}), 400
        
        # Get parameters
        threshold = float(request.form.get('threshold', DEFAULT_THRESHOLD))
        assignment_group_filter = request.form.get('assignment_group', 'All')
        
        print(f"Analysis parameters: threshold={threshold}, assignment_group={assignment_group_filter}")
        
        # Find similar tickets (now includes language filtering)
        results, lang_stats = calculate_semantic_similarity(open_df, closed_df, assignment_group_filter, threshold)
        
        # Calculate statistics
        total_open_tickets = len(open_df)
        total_closed_tickets = len(closed_df)
        
        # Use language-aware statistics
        english_open_tickets = lang_stats.get('english_open_tickets', total_open_tickets)
        non_english_open_tickets = lang_stats.get('non_english_open_tickets', 0)
        filtered_english_open_tickets = lang_stats.get('filtered_english_open_tickets', english_open_tickets)
        
        total_matches = sum(result['total_similar_closed'] for result in results)
        
        response = {
            'success': True,
            'results': results,
            'total_open_tickets': total_open_tickets,
            'english_open_tickets': english_open_tickets,
            'non_english_open_tickets': non_english_open_tickets,
            'total_closed_tickets': total_closed_tickets,
            'filtered_open_tickets': filtered_english_open_tickets,
            'open_tickets_with_matches': len(results),
            'total_matches': total_matches,
            'threshold': threshold,
            'threshold_percentage': f"{threshold * 100:.0f}%",
            'assignment_group_filter': assignment_group_filter,
            'required_columns': REQUIRED_COLUMNS,
            'preferred_columns': PREFERRED_COLUMNS,
            'open_columns': open_df.columns.tolist(),
            'closed_columns': closed_df.columns.tolist(),
            'has_closing_note': 'Close Notes' in closed_df.columns,
            'has_resolved_by': 'Resolved by' in closed_df.columns,
            'llm_available': llm is not None,
            'ai_enhanced_count': sum(1 for r in results if 'ROOT CAUSE ANALYSIS:' in r.get('root_cause', '')),
            'language_filtering_enabled': lang_stats.get('non_english_open_tickets', 0) > 0
        }
        
        return jsonify(response)
        
    except Exception as e:
        print(f"Analysis error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/language_stats')
def language_stats():
    """
    Endpoint to get language detection statistics without running full analysis
    """
    try:
        if not os.path.exists(OPEN_TICKETS_PATH):
            return jsonify({'success': False, 'error': 'Open tickets file not found'}), 404
        
        # Load open tickets
        open_df = pd.read_excel(OPEN_TICKETS_PATH)
        
        # Get language statistics
        try:
            non_english_df, english_df = split_short_descriptions(open_df)
            
            # Get language breakdown
            lang_breakdown = {}
            if len(non_english_df) > 0:
                lang_counts = non_english_df['DetectedLang'].value_counts()
                lang_breakdown = dict(lang_counts)
            
            return jsonify({
                'success': True,
                'total_tickets': len(open_df),
                'english_tickets': len(english_df),
                'non_english_tickets': len(non_english_df),
                'non_english_percentage': round((len(non_english_df) / len(open_df)) * 100, 1),
                'language_breakdown': lang_breakdown,
                'detection_available': True
            })
            
        except ImportError:
            return jsonify({
                'success': True,
                'total_tickets': len(open_df),
                'english_tickets': len(open_df),
                'non_english_tickets': 0,
                'non_english_percentage': 0,
                'language_breakdown': {},
                'detection_available': False,
                'error': 'Language detection library (langdetect) not available'
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Error in language detection: {str(e)}'
            }), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/llm_status')
def llm_status():
    """
    Check if Gemini API is available and working
    """
    global gemini_client
    
    if gemini_client is None:
        return jsonify({
            'available': False,
            'error': 'Gemini client not initialized',
            'model': None
        })
    
    try:
        # Test the Gemini client with a simple message
        is_working, test_result = gemini_client.test_connection()
        if is_working:
            return jsonify({
                'available': True,
                'status': 'Working',
                'model': 'Google Gemini 2.0 Flash',
                'test_response_length': len(test_result)
            })
        else:
            return jsonify({
                'available': False,
                'error': test_result,
                'model': 'Google Gemini 2.0 Flash'
            })
    except Exception as e:
        logging.error(f"Gemini API test failed: {str(e)}")
        return jsonify({
            'available': False,
            'error': str(e),
            'model': 'Google Gemini 2.0 Flash'
        })

@app.route('/configure_llm', methods=['POST'])
def configure_llm():
    """
    Endpoint to configure Gemini settings
    """
    global llm, gemini_client
    
    try:
        data = request.get_json()
        api_key = data.get('api_key')
        model_name = data.get('model', 'gemini-2.0-flash')
        
        # Configure Gemini client
        if api_key:
            gemini_client = GeminiClient(api_key)
        else:
            # Use existing API key from environment or default
            existing_key = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY)
            gemini_client = GeminiClient(existing_key)
        
        llm = gemini_client  # Keep compatibility
        
        # Test the Gemini client
        is_working, test_result = gemini_client.test_connection()
        
        if is_working:
            logging.info(f"Gemini client reconfigured successfully with model: {model_name}")
            
            return jsonify({
                'success': True, 
                'message': f'Google Gemini configured successfully with model: {model_name}',
                'model': model_name,
                'test_response': test_result[:100] + "..." if len(test_result) > 100 else test_result
            })
        else:
            raise Exception(f"Gemini API test failed: {test_result}")
        
    except Exception as e:
        logging.error(f"Failed to configure Gemini API: {str(e)}")
        return jsonify({
            'success': False, 
            'error': f'Failed to configure Gemini API: {str(e)}'
        }), 500

@app.route('/export_excel', methods=['POST'])
def export_excel():
    try:
        # Get the analysis results from the request
        data = request.get_json()
        
        if not data or 'results' not in data:
            return jsonify({'success': False, 'error': 'No analysis results to export'}), 400
        
        results = data['results']
        analysis_params = data.get('analysis_params', {})
        lang_stats = data.get('language_stats', None)
        
        # Create Excel file with language statistics
        excel_buffer = create_excel_export(results, analysis_params, lang_stats)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ticket_similarity_analysis_english_only_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    # Print Gemini API status on startup
    if llm is not None:
        print("✅ Google Gemini API initialized successfully - AI-enhanced suggestions available")
        try:
            # Test connection
            is_working, test_result = gemini_client.test_connection()
            if is_working:
                print("✅ Google Gemini API connection test successful")
                print(f"   Test response: {test_result[:100]}...")
            else:
                print(f"⚠️ Google Gemini API connection test failed: {test_result}")
        except Exception as e:
            print(f"⚠️ Google Gemini API connection test failed: {str(e)}")
    else:
        print("⚠️ Google Gemini API not available - will use basic suggestion format")
        print("   Please check your API key and network connection")
    
    # Print language detection status
    try:
        from langdetect import detect_langs
        print("✅ Language detection (langdetect) available - Non-English tickets will be filtered")
    except ImportError:
        print("⚠️ Language detection (langdetect) not available - All tickets will be processed")
        print("   Install with: pip install langdetect")
    
    app.run(debug=True)
